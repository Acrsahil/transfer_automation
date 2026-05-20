from pathlib import Path
import logging
import os
import time
from datetime import datetime, timedelta

import psycopg2
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv(dotenv_path=Path('.env').resolve(), override=True)

# =============================================================================
# CONFIG — edit these or set in .env
# =============================================================================

DB_HOST = os.getenv("DB_HOST")
DB_NAME = os.getenv("DB_NAME")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")
DB_PORT = os.getenv("DB_PORT", 5432)


LOCATION_NAMES = [l.strip() for l in os.getenv(
    "LOCATION_NAMES", "JS-ch/Stockz").split(",")]
ADU_PERIOD_DAYS = int(os.getenv("ADU_PERIOD_DAYS", 30))
OUTPUT_FILE = os.getenv("OUTPUT_FILE", "adu_dii_results.xlsx")
BATCH_SIZE = 500

# =============================================================================
# LOGGING
# =============================================================================
logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s")
_logger = logging.getLogger(__name__)

# =============================================================================
# DATABASE — read-only connection
# =============================================================================
DB_CONFIG = {
    "host":     os.getenv("DB_HOST", "localhost"),
    "port":     int(os.getenv("DB_PORT", 5432)),
    "dbname":   os.getenv("DB_NAME", "odoo"),
    "user":     os.getenv("DB_USER", "odoo"),
    "password": os.getenv("DB_PASSWORD", "odoo"),
}


def get_connection():
    conn = psycopg2.connect(**DB_CONFIG)
    conn.set_session(readonly=True, autocommit=True)
    return conn


# =============================================================================
# LOCATION HELPERS
# =============================================================================

def get_location_by_name(cr, complete_name: str) -> dict | None:
    cr.execute("""
        SELECT id, name, complete_name
        FROM   stock_location
        WHERE  complete_name = %s AND active = TRUE
        LIMIT  1
    """, (complete_name,))
    row = cr.fetchone()
    if not row:
        return None
    return dict(zip([d[0] for d in cr.description], row))


def get_all_child_location_ids(cr, location_id: int) -> list:
    cr.execute("""
        WITH RECURSIVE children AS (
            SELECT id FROM stock_location WHERE location_id = %s
            UNION ALL
            SELECT sl.id FROM stock_location sl
            JOIN children c ON sl.location_id = c.id
        )
        SELECT id FROM children
    """, (location_id,))
    return [r[0] for r in cr.fetchall()]


def get_dest_locations_and_picking_codes(cr, complete_name: str):
    if complete_name == "WH/Stock":
        cr.execute(
            "SELECT id FROM stock_location WHERE usage IN ('internal','customer')")
        dest_ids = tuple(r[0] for r in cr.fetchall()) or (None,)
        picking_codes = ("outgoing", "internal")
    else:
        cr.execute(
            "SELECT id FROM stock_location WHERE usage = 'customer' LIMIT 1")
        row = cr.fetchone()
        dest_ids = (row[0],) if row else (None,)
        picking_codes = ("outgoing",)
    return dest_ids, picking_codes


def get_picking_type_ids(cr, codes: tuple) -> tuple:
    cr.execute(
        "SELECT id FROM stock_picking_type WHERE code = ANY(%s)", (list(codes),))
    return tuple(r[0] for r in cr.fetchall()) or (None,)


# =============================================================================
# STOCK STATUS
# =============================================================================

def stock_status(dii):
    if dii is None or dii <= 0:
        return "no_sales"
    if dii < 7:
        return "critical"
    if dii < 14:
        return "low"
    if dii <= 60:
        return "healthy"
    return "overstock"


# =============================================================================
# STATUS COLOR MAP
# =============================================================================

STATUS_FILLS = {
    "critical":  PatternFill("solid", start_color="FFCCCC"),   # light red
    "low":       PatternFill("solid", start_color="FFE5CC"),   # light orange
    "healthy":   PatternFill("solid", start_color="CCFFCC"),   # light green
    "overstock": PatternFill("solid", start_color="CCE5FF"),   # light blue
    "no_sales":  PatternFill("solid", start_color="F2F2F2"),   # light grey
}

# =============================================================================
# CORE CALCULATION QUERY  (pure SELECT, no writes)
# =============================================================================

ADU_CALC_SQL = """
WITH sales_data AS (
    SELECT
        sml.product_id,
        SUM(sml.qty_done) AS total_sold
    FROM  stock_move_line sml
    INNER JOIN stock_picking   sp ON sp.id = sml.picking_id
    INNER JOIN product_product pp ON pp.id = sml.product_id
    WHERE sp.picking_type_id       IN %(outgoing_picking_type_ids)s
      AND sp.location_dest_id      IN %(dest_location_ids)s
      AND sml.location_id          IN %(location_ids)s
      AND sml.location_dest_id NOT IN %(location_ids)s
      AND sml.location_dest_id     IN %(dest_location_ids)s
      AND sml.state                 = 'done'
      AND sml.date                 >= %(start_date)s
      AND sml.date                 <= %(end_date)s
      AND pp.active                 = TRUE
      AND COALESCE(pp.is_ondemand_product, FALSE) = FALSE
      AND pp.id                    >= %(min_prod_id)s
      AND pp.id                     < %(max_prod_id)s
    GROUP BY sml.product_id
),
stock_data AS (
    SELECT
        sq.product_id,
        SUM(sq.quantity)                             AS on_hand_qty,
        SUM(sq.reserved_quantity)                    AS reserved_qty,
        SUM(sq.quantity) - SUM(sq.reserved_quantity) AS available_qty
    FROM  stock_quant       sq
    INNER JOIN product_product pp ON pp.id = sq.product_id
    WHERE sq.location_id                             IN %(location_ids)s
      AND pp.active                                   = TRUE
      AND COALESCE(pp.is_ondemand_product, FALSE)     = FALSE
      AND pp.id                                      >= %(min_prod_id)s
      AND pp.id                                       < %(max_prod_id)s
    GROUP BY sq.product_id
),
calendar AS (
    SELECT generate_series(
        %(start_date)s::date,
        %(end_date)s::date,
        INTERVAL '1 day'
    )::date AS day
),
daily_moves AS (
    SELECT
        m.product_id,
        m.day,
        SUM(m.qty)                                      AS net_qty,
        SUM(CASE WHEN m.qty > 0 THEN m.qty ELSE 0 END) AS in_qty
    FROM (
        SELECT sml.product_id, sml.date::date AS day, sml.qty_done AS qty
        FROM   stock_move_line sml
        INNER JOIN product_product pp ON pp.id = sml.product_id
        WHERE  sml.state              = 'done'
          AND  sml.location_dest_id  IN %(location_ids)s
          AND  sml.date              >= %(start_date)s
          AND  sml.date              <= %(end_date)s
          AND  pp.active              = TRUE
          AND  COALESCE(pp.is_ondemand_product, FALSE) = FALSE
          AND  pp.id                 >= %(min_prod_id)s
          AND  pp.id                  < %(max_prod_id)s
        UNION ALL
        SELECT sml.product_id, sml.date::date AS day, -sml.qty_done AS qty
        FROM   stock_move_line sml
        INNER JOIN product_product pp ON pp.id = sml.product_id
        WHERE  sml.state          = 'done'
          AND  sml.location_id   IN %(location_ids)s
          AND  sml.date          >= %(start_date)s
          AND  sml.date          <= %(end_date)s
          AND  pp.active          = TRUE
          AND  COALESCE(pp.is_ondemand_product, FALSE) = FALSE
          AND  pp.id             >= %(min_prod_id)s
          AND  pp.id              < %(max_prod_id)s
    ) m
    GROUP BY m.product_id, m.day
),
opening_stock AS (
    SELECT
        sml.product_id,
        SUM(CASE
            WHEN sml.location_dest_id IN %(location_ids)s THEN  sml.qty_done
            WHEN sml.location_id      IN %(location_ids)s THEN -sml.qty_done
            ELSE 0
        END) AS opening_qty
    FROM stock_move_line sml
    INNER JOIN product_product pp ON pp.id = sml.product_id
    WHERE sml.state = 'done'
      AND sml.date  < %(start_date)s
      AND (sml.location_id IN %(location_ids)s OR sml.location_dest_id IN %(location_ids)s)
      AND pp.active  = TRUE
      AND COALESCE(pp.is_ondemand_product, FALSE) = FALSE
      AND pp.id     >= %(min_prod_id)s
      AND pp.id      < %(max_prod_id)s
    GROUP BY sml.product_id
),
daily_balance AS (
    SELECT
        pp.id AS product_id,
        c.day,
        COALESCE(dm.net_qty, 0) AS daily_net_qty,
        COALESCE(dm.in_qty,  0) AS daily_in_qty,
        COALESCE(os.opening_qty, 0)
        + COALESCE(
            SUM(dm.net_qty) OVER (
                PARTITION BY pp.id ORDER BY c.day
                ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
            ), 0
        ) AS running_qty
    FROM product_product pp
    CROSS JOIN calendar c
    LEFT JOIN opening_stock os ON os.product_id = pp.id
    LEFT JOIN daily_moves   dm ON dm.product_id = pp.id AND dm.day = c.day
    WHERE pp.active = TRUE
      AND COALESCE(pp.is_ondemand_product, FALSE) = FALSE
      AND pp.id >= %(min_prod_id)s
      AND pp.id  < %(max_prod_id)s
),
active_days_data AS (
    SELECT
        product_id,
        COUNT(*) FILTER (
            WHERE (running_qty - daily_net_qty) > 0 OR daily_in_qty > 0
        ) AS active_days
    FROM daily_balance
    GROUP BY product_id
),
product_names AS (
    SELECT pp.id, pt.name AS product_name, pp.default_code AS internal_ref
    FROM   product_product pp
    JOIN   product_template pt ON pt.id = pp.product_tmpl_id
    WHERE  pp.active = TRUE
      AND  COALESCE(pp.is_ondemand_product, FALSE) = FALSE
      AND  pp.id >= %(min_prod_id)s
      AND  pp.id  < %(max_prod_id)s
)
SELECT
    COALESCE(sd.product_id, st.product_id)    AS product_id,
    pn.internal_ref,
    pn.product_name,
    COALESCE(st.available_qty,  0.0)           AS current_stock_qty,
    COALESCE(st.on_hand_qty,    0.0)           AS on_hand_qty,
    COALESCE(st.reserved_qty,   0.0)           AS reserved_qty,
    COALESCE(sd.total_sold,     0.0)           AS total_sold,
    COALESCE(ad.active_days, %(period_days)s)  AS active_days,
    %(period_days)s                            AS calculation_period,
    CASE
        WHEN COALESCE(ad.active_days, 0) > 0
        THEN ROUND((COALESCE(sd.total_sold, 0) / ad.active_days)::numeric, 3)
        ELSE 0
    END AS adu_value,
    CASE
        WHEN COALESCE(sd.total_sold, 0.0) > 0
         AND COALESCE(ad.active_days, 0)   > 0
        THEN ROUND(
                (COALESCE(st.available_qty, 0.0) / (sd.total_sold / ad.active_days))::numeric,
                2
             )
        ELSE NULL
    END AS dii_value
FROM      sales_data sd
FULL OUTER JOIN stock_data      st ON sd.product_id = st.product_id
LEFT JOIN  active_days_data     ad ON COALESCE(sd.product_id, st.product_id) = ad.product_id
LEFT JOIN  product_names        pn ON COALESCE(sd.product_id, st.product_id) = pn.id
ORDER BY adu_value DESC NULLS LAST
"""

CSV_COLUMNS = [
    "location_id", "location_name",
    "product_id", "internal_ref", "product_name",
    "on_hand_qty", "reserved_qty", "current_stock_qty",
    "total_sold", "active_days", "calculation_period",
    "adu_value", "dii_value", "stock_status",
    "calculated_at",
]

HEADERS = [
    "Location ID", "Location Name",
    "Product ID", "Internal Ref", "Product Name",
    "On Hand Qty", "Reserved Qty", "Available Qty",
    "Total Sold", "Active Days", "Calc Period (days)",
    "ADU Value", "DII Value", "Stock Status",
    "Calculated At",
]

HEADER_FILL = PatternFill("solid", start_color="2F4F8F")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DATA_FONT = Font(name="Arial", size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left",   vertical="center")

COL_WIDTHS = [12, 20, 12, 14, 40, 12, 12, 12, 12, 12, 18, 12, 12, 14, 20]


# =============================================================================
# EXCEL WRITER — initialise workbook once, append per batch
# =============================================================================

class ExcelWriter:
    def __init__(self, output_file: str):
        self.output_file = output_file
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "ADU DII Results"
        self._write_header()
        self._row = 2  # next data row

    def _write_header(self):
        for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
            cell = self.ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            self.ws.column_dimensions[get_column_letter(col_idx)].width = width
        self.ws.freeze_panes = "A2"

    @staticmethod
    def _to_scalar(value):
        """Flatten Odoo translated-field dicts to a plain string."""
        if isinstance(value, dict):
            return value.get("en_US") or next(iter(value.values()), "")
        return value

    def append_rows(self, rows: list):
        for row in rows:
            status = stock_status(row["dii_value"])
            row["stock_status"] = status
            row["calculated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            fill = STATUS_FILLS.get(status)

            values = [self._to_scalar(row.get(col)) for col in CSV_COLUMNS]
            for col_idx, value in enumerate(values, 1):
                cell = self.ws.cell(row=self._row, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.alignment = LEFT_ALIGN
                if fill:
                    cell.fill = fill
            self._row += 1

        # Save after every batch so progress is always on disk
        self.wb.save(self.output_file)

    def total_rows(self) -> int:
        return self._row - 2


# =============================================================================
# CORE CALCULATION QUERY RUNNER
# =============================================================================

def calculate_batch(cr, location_ids, dest_location_ids, start_date, end_date,
                    outgoing_picking_type_ids, period_days, min_prod_id, max_prod_id) -> list:
    cr.execute(ADU_CALC_SQL, {
        "location_ids":              location_ids,
        "dest_location_ids":         dest_location_ids,
        "start_date":                start_date,
        "end_date":                  end_date,
        "outgoing_picking_type_ids": outgoing_picking_type_ids,
        "period_days":               period_days,
        "min_prod_id":               min_prod_id,
        "max_prod_id":               max_prod_id,
    })
    col_names = [d[0] for d in cr.description]
    return [dict(zip(col_names, row)) for row in cr.fetchall()]


# =============================================================================
# PER-LOCATION ORCHESTRATOR  — writes to Excel after each batch
# =============================================================================

def calculate_for_location(cr, location: dict, period_days: int, writer: ExcelWriter) -> int:
    loc_name = location["complete_name"]
    _logger.info("Processing: %s (id=%s)", loc_name, location["id"])
    t0 = time.time()

    dest_location_ids, picking_codes = get_dest_locations_and_picking_codes(
        cr, loc_name)
    outgoing_picking_type_ids = get_picking_type_ids(cr, picking_codes)
    child_ids = get_all_child_location_ids(cr, location["id"])
    location_ids = tuple([location["id"]] + child_ids)

    end_date = datetime.now()
    start_date = end_date - timedelta(days=period_days - 1)

    _logger.info(
        "  %d location ID(s) | %s to %s | picking=%s",
        len(location_ids), start_date.date(), end_date.date(), picking_codes
    )

    cr.execute("""
        SELECT MIN(id), MAX(id) FROM product_product
        WHERE active = TRUE AND COALESCE(is_ondemand_product, FALSE) = FALSE
    """)
    min_id, max_id = cr.fetchone()
    if min_id is None:
        _logger.info("  No active products — skipping.")
        return 0

    total_batches = (max_id - min_id + BATCH_SIZE) // BATCH_SIZE
    written = 0

    for i, batch_min in enumerate(range(min_id, max_id + 1, BATCH_SIZE), 1):
        batch_max = batch_min + BATCH_SIZE
        _logger.info("  Batch %d/%d — product IDs %d to %d",
                     i, total_batches, batch_min, batch_max)

        rows = calculate_batch(
            cr, location_ids, dest_location_ids, start_date, end_date,
            outgoing_picking_type_ids, period_days, batch_min, batch_max
        )

        # ── Enrich rows with location metadata before writing ──
        for row in rows:
            row["location_id"] = location["id"]
            row["location_name"] = loc_name

        # ── Write this batch to Excel immediately ──────────────
        writer.append_rows(rows)
        written += len(rows)
        _logger.info("  → %d rows written to Excel (total so far: %d)", len(
            rows), writer.total_rows())

    _logger.info("  %d products | %.0f ms", written, (time.time() - t0) * 1000)
    return written


# =============================================================================
# MAIN
# =============================================================================

def run():
    _logger.info(
        "Connecting (read-only): %(dbname)s @ %(host)s:%(port)s", DB_CONFIG)
    _logger.info("Locations : %s", LOCATION_NAMES)
    _logger.info("ADU period: %d days", ADU_PERIOD_DAYS)
    _logger.info("Output    : %s", OUTPUT_FILE)

    writer = ExcelWriter(OUTPUT_FILE)
    conn = get_connection()

    try:
        with conn.cursor() as cr:
            for loc_name in LOCATION_NAMES:
                location = get_location_by_name(cr, loc_name)
                if not location:
                    _logger.warning(
                        "Location not found: '%s' — skipping.", loc_name)
                    continue
                calculate_for_location(cr, location, ADU_PERIOD_DAYS, writer)
    finally:
        conn.close()

    total = writer.total_rows()
    if total == 0:
        _logger.warning("No results written.")
        return

    _logger.info("Results written to: %s (%d rows)", OUTPUT_FILE, total)

    # Count statuses from sheet (no second pass needed — just read column 14)
    counts = {"critical": 0, "low": 0,
              "healthy": 0, "overstock": 0, "no_sales": 0}
    for row in writer.ws.iter_rows(min_row=2, max_row=writer.ws.max_row, min_col=14, max_col=14):
        status = row[0].value
        if status in counts:
            counts[status] += 1

    _logger.info("=" * 50)
    _logger.info("SUMMARY")
    _logger.info("=" * 50)
    _logger.info("  Total products : %d", total)
    _logger.info("  Critical (<7d) : %d", counts["critical"])
    _logger.info("  Low (7-14d)    : %d", counts["low"])
    _logger.info("  Healthy        : %d", counts["healthy"])
    _logger.info("  Overstock      : %d", counts["overstock"])
    _logger.info("  No sales       : %d", counts["no_sales"])
    _logger.info("=" * 50)


if __name__ == "__main__":
    run()
