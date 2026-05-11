from pathlib import Path
import psycopg2
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
from dotenv import load_dotenv

load_dotenv(dotenv_path=Path('.env').resolve(), override=True)

# ─────────────────────────────────────────────
# 1. DATABASE CONNECTION
# ─────────────────────────────────────────────
DB_CONFIG = {
    'host':     os.getenv('DB_HOST', 'localhost'),
    'port':     int(os.getenv('DB_PORT', 5432)),
    'dbname':   os.getenv('DB_NAME', 'odoo'),
    'user':     os.getenv('DB_USER', 'odoo'),
    'password': os.getenv('DB_PASSWORD', ''),
}

# ─────────────────────────────────────────────
# 2. TEST PARAMETERS
# ─────────────────────────────────────────────
PERIOD_DAYS = 7
LOCATION_IDS = (5,)
LOCATION_ID = 5
DEST_LOCATION_IDS = (8,)
PICKING_TYPE_IDS = (3,)
MIN_PROD_ID = 1
MAX_PROD_ID = 9999999

END_DATE = datetime.now()
START_DATE = END_DATE - timedelta(days=PERIOD_DAYS - 1)

PARAMS = {
    'location_ids':             LOCATION_IDS,
    'location_id':              LOCATION_ID,
    'dest_location_ids':        DEST_LOCATION_IDS,
    'outgoing_picking_type_id': PICKING_TYPE_IDS,
    'start_date':               START_DATE,
    'end_date':                 END_DATE,
    'period_days':              PERIOD_DAYS,
    'min_prod_id':              MIN_PROD_ID,
    'max_prod_id':              MAX_PROD_ID,
}

OUTPUT_FILE = 'adu_dii_results.xlsx'

# ─────────────────────────────────────────────
# 3. CTE QUERIES
# ─────────────────────────────────────────────
QUERIES = {
    'sales_data': """
        SELECT
            sml.product_id,
            pt.name                         AS product_name,
            pp.default_code                 AS internal_ref,
            SUM(sml.qty_done)               AS total_sold
        FROM stock_move_line sml
        INNER JOIN stock_picking sp         ON sp.id = sml.picking_id
        INNER JOIN product_product pp       ON sml.product_id = pp.id
        INNER JOIN product_template pt      ON pt.id = pp.product_tmpl_id
        WHERE sp.picking_type_id IN %(outgoing_picking_type_id)s
            AND sp.location_dest_id IN %(dest_location_ids)s
            AND sml.location_id IN %(location_ids)s
            AND sml.location_dest_id NOT IN %(location_ids)s
            AND sml.location_dest_id IN %(dest_location_ids)s
            AND sml.state = 'done'
            AND sml.date >= %(start_date)s AND sml.date <= %(end_date)s
            AND pp.active = TRUE AND COALESCE(pp.is_ondemand_product, FALSE) = FALSE
            AND pp.id >= %(min_prod_id)s AND pp.id < %(max_prod_id)s
        GROUP BY sml.product_id, pt.name, pp.default_code
        ORDER BY total_sold DESC
    """,

    'stock_data': """
        SELECT
            sq.product_id,
            pt.name                                         AS product_name,
            pp.default_code                                 AS internal_ref,
            SUM(sq.quantity)                                AS on_hand_qty,
            SUM(sq.reserved_quantity)                       AS reserved_qty,
            SUM(sq.quantity) - SUM(sq.reserved_quantity)    AS available_qty
        FROM stock_quant sq
        INNER JOIN product_product pp   ON sq.product_id = pp.id
        INNER JOIN product_template pt  ON pt.id = pp.product_tmpl_id
        WHERE sq.location_id IN %(location_ids)s
            AND pp.active = TRUE AND COALESCE(pp.is_ondemand_product, FALSE) = FALSE
            AND pp.id >= %(min_prod_id)s AND pp.id < %(max_prod_id)s
        GROUP BY sq.product_id, pt.name, pp.default_code
        ORDER BY on_hand_qty DESC
    """,

    'calendar': """
        SELECT generate_series(
            %(start_date)s::date,
            %(end_date)s::date,
            INTERVAL '1 day'
        )::date AS day
    """,

    'daily_moves': """
        SELECT
            m.product_id,
            pt.name                                         AS product_name,
            pp.default_code                                 AS internal_ref,
            m.day,
            SUM(m.qty)                                      AS net_qty,
            SUM(CASE WHEN m.qty > 0 THEN m.qty ELSE 0 END) AS in_qty
        FROM (
            SELECT sml.product_id, sml.date::date AS day, sml.qty_done AS qty
            FROM stock_move_line sml
            INNER JOIN product_product pp ON sml.product_id = pp.id
            WHERE sml.state = 'done'
                AND sml.location_dest_id IN %(location_ids)s
                AND sml.date >= %(start_date)s AND sml.date <= %(end_date)s
                AND pp.active = TRUE AND COALESCE(pp.is_ondemand_product, FALSE) = FALSE
                AND pp.id >= %(min_prod_id)s AND pp.id < %(max_prod_id)s
            UNION ALL
            SELECT sml.product_id, sml.date::date AS day, -sml.qty_done AS qty
            FROM stock_move_line sml
            INNER JOIN product_product pp ON sml.product_id = pp.id
            WHERE sml.state = 'done'
                AND sml.location_id IN %(location_ids)s
                AND sml.date >= %(start_date)s AND sml.date <= %(end_date)s
                AND pp.active = TRUE AND COALESCE(pp.is_ondemand_product, FALSE) = FALSE
                AND pp.id >= %(min_prod_id)s AND pp.id < %(max_prod_id)s
        ) m
        INNER JOIN product_product pp   ON pp.id = m.product_id
        INNER JOIN product_template pt  ON pt.id = pp.product_tmpl_id
        GROUP BY m.product_id, pt.name, pp.default_code, m.day
        ORDER BY m.product_id, m.day
    """,

    'opening_stock': """
        SELECT
            sml.product_id,
            pt.name                 AS product_name,
            pp.default_code         AS internal_ref,
            SUM(CASE
                WHEN sml.location_dest_id IN %(location_ids)s THEN  sml.qty_done
                WHEN sml.location_id      IN %(location_ids)s THEN -sml.qty_done
                ELSE 0
            END) AS opening_qty
        FROM stock_move_line sml
        INNER JOIN product_product pp   ON sml.product_id = pp.id
        INNER JOIN product_template pt  ON pt.id = pp.product_tmpl_id
        WHERE sml.state = 'done'
            AND sml.date < %(start_date)s
            AND (sml.location_id IN %(location_ids)s OR sml.location_dest_id IN %(location_ids)s)
            AND pp.active = TRUE AND COALESCE(pp.is_ondemand_product, FALSE) = FALSE
            AND pp.id >= %(min_prod_id)s AND pp.id < %(max_prod_id)s
        GROUP BY sml.product_id, pt.name, pp.default_code
        ORDER BY opening_qty DESC
    """,

    'active_days': """
        WITH calendar AS (
            SELECT generate_series(%(start_date)s::date, %(end_date)s::date, INTERVAL '1 day')::date AS day
        ),
        daily_moves AS (
            SELECT m.product_id, m.day,
                SUM(m.qty) AS net_qty,
                SUM(CASE WHEN m.qty > 0 THEN m.qty ELSE 0 END) AS in_qty
            FROM (
                SELECT sml.product_id, sml.date::date AS day, sml.qty_done AS qty
                FROM stock_move_line sml INNER JOIN product_product pp ON sml.product_id = pp.id
                WHERE sml.state='done' AND sml.location_dest_id IN %(location_ids)s
                    AND sml.date >= %(start_date)s AND sml.date <= %(end_date)s
                    AND pp.active=TRUE AND COALESCE(pp.is_ondemand_product,FALSE)=FALSE
                    AND pp.id >= %(min_prod_id)s AND pp.id < %(max_prod_id)s
                UNION ALL
                SELECT sml.product_id, sml.date::date AS day, -sml.qty_done AS qty
                FROM stock_move_line sml INNER JOIN product_product pp ON sml.product_id = pp.id
                WHERE sml.state='done' AND sml.location_id IN %(location_ids)s
                    AND sml.date >= %(start_date)s AND sml.date <= %(end_date)s
                    AND pp.active=TRUE AND COALESCE(pp.is_ondemand_product,FALSE)=FALSE
                    AND pp.id >= %(min_prod_id)s AND pp.id < %(max_prod_id)s
            ) m GROUP BY m.product_id, m.day
        ),
        opening_stock AS (
            SELECT sml.product_id,
                SUM(CASE
                    WHEN sml.location_dest_id IN %(location_ids)s THEN  sml.qty_done
                    WHEN sml.location_id      IN %(location_ids)s THEN -sml.qty_done
                    ELSE 0 END) AS opening_qty
            FROM stock_move_line sml INNER JOIN product_product pp ON sml.product_id = pp.id
            WHERE sml.state='done' AND sml.date < %(start_date)s
                AND (sml.location_id IN %(location_ids)s OR sml.location_dest_id IN %(location_ids)s)
                AND pp.active=TRUE AND COALESCE(pp.is_ondemand_product,FALSE)=FALSE
                AND pp.id >= %(min_prod_id)s AND pp.id < %(max_prod_id)s
            GROUP BY sml.product_id
        ),
        daily_balance AS (
            SELECT pp.id AS product_id, c.day,
                COALESCE(dm.net_qty,0) AS daily_net_qty,
                COALESCE(dm.in_qty,0)  AS daily_in_qty,
                COALESCE(os.opening_qty,0)
                + COALESCE(SUM(dm.net_qty) OVER (
                    PARTITION BY pp.id ORDER BY c.day ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
                ),0) AS running_qty
            FROM product_product pp
            CROSS JOIN calendar c
            LEFT JOIN opening_stock os ON os.product_id = pp.id
            LEFT JOIN daily_moves   dm ON dm.product_id = pp.id AND dm.day = c.day
            WHERE pp.active=TRUE AND COALESCE(pp.is_ondemand_product,FALSE)=FALSE
                AND pp.id >= %(min_prod_id)s AND pp.id < %(max_prod_id)s
        )
        SELECT
            db.product_id,
            pt.name             AS product_name,
            pp.default_code     AS internal_ref,
            COUNT(*) FILTER (WHERE (running_qty - daily_net_qty) > 0 OR daily_in_qty > 0) AS active_days
        FROM daily_balance db
        INNER JOIN product_product pp   ON pp.id = db.product_id
        INNER JOIN product_template pt  ON pt.id = pp.product_tmpl_id
        GROUP BY db.product_id, pt.name, pp.default_code
        ORDER BY active_days DESC
    """,

    'calculated_ADU_DII': """
        WITH sales_data AS (
            SELECT sml.product_id, SUM(sml.qty_done) AS total_sold
            FROM stock_move_line sml
            INNER JOIN stock_picking sp ON sp.id = sml.picking_id
            INNER JOIN product_product pp ON sml.product_id = pp.id
            WHERE sp.picking_type_id IN %(outgoing_picking_type_id)s
                AND sp.location_dest_id IN %(dest_location_ids)s
                AND sml.location_id IN %(location_ids)s
                AND sml.location_dest_id NOT IN %(location_ids)s
                AND sml.location_dest_id IN %(dest_location_ids)s
                AND sml.state = 'done'
                AND sml.date >= %(start_date)s AND sml.date <= %(end_date)s
                AND pp.active=TRUE AND COALESCE(pp.is_ondemand_product,FALSE)=FALSE
                AND pp.id >= %(min_prod_id)s AND pp.id < %(max_prod_id)s
            GROUP BY sml.product_id
        ),
        stock_data AS (
            SELECT sq.product_id,
                SUM(sq.quantity) AS on_hand_qty,
                SUM(sq.reserved_quantity) AS reserved_qty,
                SUM(sq.quantity) - SUM(sq.reserved_quantity) AS available_qty
            FROM stock_quant sq
            INNER JOIN product_product pp ON sq.product_id = pp.id
            WHERE sq.location_id IN %(location_ids)s
                AND pp.active=TRUE AND COALESCE(pp.is_ondemand_product,FALSE)=FALSE
                AND pp.id >= %(min_prod_id)s AND pp.id < %(max_prod_id)s
            GROUP BY sq.product_id
        ),
        calendar AS (
            SELECT generate_series(%(start_date)s::date, %(end_date)s::date, INTERVAL '1 day')::date AS day
        ),
        daily_moves AS (
            SELECT m.product_id, m.day,
                SUM(m.qty) AS net_qty,
                SUM(CASE WHEN m.qty > 0 THEN m.qty ELSE 0 END) AS in_qty
            FROM (
                SELECT sml.product_id, sml.date::date AS day, sml.qty_done AS qty
                FROM stock_move_line sml INNER JOIN product_product pp ON sml.product_id = pp.id
                WHERE sml.state='done' AND sml.location_dest_id IN %(location_ids)s
                    AND sml.date >= %(start_date)s AND sml.date <= %(end_date)s
                    AND pp.active=TRUE AND COALESCE(pp.is_ondemand_product,FALSE)=FALSE
                    AND pp.id >= %(min_prod_id)s AND pp.id < %(max_prod_id)s
                UNION ALL
                SELECT sml.product_id, sml.date::date AS day, -sml.qty_done AS qty
                FROM stock_move_line sml INNER JOIN product_product pp ON sml.product_id = pp.id
                WHERE sml.state='done' AND sml.location_id IN %(location_ids)s
                    AND sml.date >= %(start_date)s AND sml.date <= %(end_date)s
                    AND pp.active=TRUE AND COALESCE(pp.is_ondemand_product,FALSE)=FALSE
                    AND pp.id >= %(min_prod_id)s AND pp.id < %(max_prod_id)s
            ) m GROUP BY m.product_id, m.day
        ),
        opening_stock AS (
            SELECT sml.product_id,
                SUM(CASE
                    WHEN sml.location_dest_id IN %(location_ids)s THEN  sml.qty_done
                    WHEN sml.location_id      IN %(location_ids)s THEN -sml.qty_done
                    ELSE 0 END) AS opening_qty
            FROM stock_move_line sml INNER JOIN product_product pp ON sml.product_id = pp.id
            WHERE sml.state='done' AND sml.date < %(start_date)s
                AND (sml.location_id IN %(location_ids)s OR sml.location_dest_id IN %(location_ids)s)
                AND pp.active=TRUE AND COALESCE(pp.is_ondemand_product,FALSE)=FALSE
                AND pp.id >= %(min_prod_id)s AND pp.id < %(max_prod_id)s
            GROUP BY sml.product_id
        ),
        daily_balance AS (
            SELECT pp.id AS product_id, c.day,
                COALESCE(dm.net_qty,0) AS daily_net_qty,
                COALESCE(dm.in_qty,0)  AS daily_in_qty,
                COALESCE(os.opening_qty,0)
                + COALESCE(SUM(dm.net_qty) OVER (
                    PARTITION BY pp.id ORDER BY c.day ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
                ),0) AS running_qty
            FROM product_product pp
            CROSS JOIN calendar c
            LEFT JOIN opening_stock os ON os.product_id = pp.id
            LEFT JOIN daily_moves   dm ON dm.product_id = pp.id AND dm.day = c.day
            WHERE pp.active=TRUE AND COALESCE(pp.is_ondemand_product,FALSE)=FALSE
                AND pp.id >= %(min_prod_id)s AND pp.id < %(max_prod_id)s
        ),
        active_days_data AS (
            SELECT product_id,
                COUNT(*) FILTER (WHERE (running_qty - daily_net_qty) > 0 OR daily_in_qty > 0) AS active_days
            FROM daily_balance GROUP BY product_id
        )
        SELECT
            %(location_id)s                                AS location_id,
            COALESCE(sd.product_id, st.product_id)        AS product_id,
            pt.name                                        AS product_name,
            pp2.default_code                               AS internal_ref,
            COALESCE(st.available_qty, 0.0)               AS current_stock_qty,
            COALESCE(sd.total_sold,    0.0)               AS total_sold,
            COALESCE(ad.active_days, %(period_days)s)     AS active_days,
            CASE WHEN COALESCE(ad.active_days,0) > 0
                 THEN ROUND((sd.total_sold / ad.active_days)::numeric, 3)
                 ELSE 0 END                               AS adu_value,
            CASE WHEN COALESCE(sd.total_sold,0.0) > 0 AND COALESCE(ad.active_days,0) > 0
                 THEN ROUND((COALESCE(st.available_qty,0.0) / (sd.total_sold / ad.active_days))::numeric, 2)
                 ELSE NULL END                            AS dii_value,
            %(period_days)s                               AS calculation_period
        FROM sales_data sd
        FULL OUTER JOIN stock_data st       ON sd.product_id = st.product_id
        LEFT JOIN active_days_data ad       ON COALESCE(sd.product_id, st.product_id) = ad.product_id
        LEFT JOIN product_product pp2       ON pp2.id = COALESCE(sd.product_id, st.product_id)
        LEFT JOIN product_template pt       ON pt.id = pp2.product_tmpl_id
        ORDER BY adu_value DESC NULLS LAST
    """,
}

# ─────────────────────────────────────────────
# 4. EXCEL HELPERS
# ─────────────────────────────────────────────
HEADER_FILL = PatternFill('solid', start_color='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
DATA_FONT = Font(name='Arial', size=10)
ALT_FILL = PatternFill('solid', start_color='D6E4F0')
CENTER = Alignment(horizontal='center', vertical='center')


def sanitize(value):
    """Odoo translated fields return {'en_US': 'Name'} — extract plain string."""
    if isinstance(value, dict):
        return value.get('en_US') or next(iter(value.values()), '')
    return value


def write_sheet(wb, sheet_name, columns, rows):
    ws = wb.create_sheet(title=sheet_name[:31])

    # Header row
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper())
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    # Sanitize every value (handles Odoo translated field dicts)
    clean_rows = [[sanitize(v) for v in row] for row in rows]

    # Data rows
    for row_idx, row in enumerate(clean_rows, start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            if fill:
                cell.fill = fill

    # Auto column width
    for col_idx, col_name in enumerate(columns, start=1):
        max_len = max(
            len(str(col_name)),
            *(len(str(row[col_idx - 1])) for row in clean_rows) if clean_rows else [0]
        )
        ws.column_dimensions[get_column_letter(
            col_idx)].width = min(max_len + 4, 40)

    # Freeze header
    ws.freeze_panes = 'A2'
    return ws


def write_summary_sheet(wb, summary):
    ws = wb.create_sheet(title='Summary', index=0)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30

    title_font = Font(bold=True, name='Arial', size=14, color='1F4E79')
    label_font = Font(bold=True, name='Arial', size=10)
    normal_font = Font(name='Arial', size=10)

    ws['A1'] = 'ADU / DII Debug Report'
    ws['A1'].font = title_font
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = normal_font

    ws['A4'] = 'Parameter'
    ws['A4'].font = label_font
    ws['B4'] = 'Value'
    ws['B4'].font = label_font
    params_display = [
        ('Period Days',     PERIOD_DAYS),
        ('Start Date',      START_DATE.strftime('%Y-%m-%d')),
        ('End Date',        END_DATE.strftime('%Y-%m-%d')),
        ('Location IDs',    str(LOCATION_IDS)),
        ('Dest Loc IDs',    str(DEST_LOCATION_IDS)),
        ('Picking Type IDs', str(PICKING_TYPE_IDS)),
        ('Min Product ID',  MIN_PROD_ID),
        ('Max Product ID',  MAX_PROD_ID),
    ]
    for i, (k, v) in enumerate(params_display, start=5):
        ws.cell(row=i, column=1, value=k).font = normal_font
        ws.cell(row=i, column=2, value=v).font = normal_font

    ws['A14'] = 'CTE / Sheet'
    ws['A14'].font = label_font
    ws['B14'] = 'Row Count'
    ws['B14'].font = label_font
    ws['C14'] = 'Status'
    ws['C14'].font = label_font
    for i, (name, count, status) in enumerate(summary, start=15):
        ws.cell(row=i, column=1, value=name).font = normal_font
        ws.cell(row=i, column=2, value=count).font = normal_font
        color = '00B050' if status == 'OK' else 'FF0000'
        cell = ws.cell(row=i, column=3, value=status)
        cell.font = Font(name='Arial', size=10, color=color, bold=True)

    ws.freeze_panes = 'A2'


# ─────────────────────────────────────────────
# 5. MAIN
# ─────────────────────────────────────────────
def main():
    print(f"\nConnecting to {DB_CONFIG['dbname']} on {DB_CONFIG['host']}...")
    conn = psycopg2.connect(**DB_CONFIG)
    conn.set_session(readonly=True, autocommit=True)
    cur = conn.cursor()

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet
    summary = []

    for sheet_name, sql in QUERIES.items():
        print(f"  Running: {sheet_name} ...", end=' ', flush=True)
        try:
            cur.execute(sql, PARAMS)
            rows = cur.fetchall()
            columns = [d[0] for d in cur.description]
            write_sheet(wb, sheet_name, columns, rows)
            summary.append((sheet_name, len(rows), 'OK'))
            print(f"{len(rows)} rows")
        except Exception as e:
            summary.append((sheet_name, 0, f'ERROR: {e}'))
            print(f"ERROR — {e}")

    write_summary_sheet(wb, summary)

    cur.close()
    conn.close()

    wb.save(OUTPUT_FILE)
    print(f"\n✅  Saved to: {OUTPUT_FILE}")
    print("    Sheets:", [s[0] for s in summary])


if __name__ == '__main__':
    main()
